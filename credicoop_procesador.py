"""
Procesador de extractos Banco Credicoop
========================================
Lee el .xls "Saldos y Movimientos" y genera un .xlsx simplificado con:
    Fecha | Descripción | Importe | Tipo

Uso:
    python credicoop_procesador.py                  # busca el primer .xls en la carpeta
    python credicoop_procesador.py archivo.xls      # archivo explícito

Requisitos:
    pip install xlrd openpyxl pandas
"""

import sys
import os
import glob
import unicodedata
import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# ── Helpers ────────────────────────────────────────────────────────────────────

def norm(s):
    """Minúsculas y sin tildes para comparar nombres de columnas."""
    s = unicodedata.normalize("NFD", str(s).lower().strip())
    return "".join(c for c in s if not unicodedata.combining(c))


def to_float(v):
    """Convierte cualquier valor (incluyendo formatos argentinos) a float."""
    if v is None:
        return 0.0
    if isinstance(v, float) and pd.isna(v):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace("\xa0", "").replace(" ", "")
    if s in ("", "nan", "NaN", "None", "-"):
        return 0.0
    # Formato argentino: 1.234,56  (punto=miles, coma=decimal)
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def fmt_fecha(v):
    """Devuelve la fecha como dd/mm/aaaa."""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    s = str(v).strip()
    if s in ("", "nan", "NaT", "None"):
        return ""
    # "2026-05-29 00:00:00" → tomar solo la parte de fecha
    if " " in s:
        s = s.split(" ")[0]
    # yyyy-mm-dd
    if "-" in s:
        try:
            dt = datetime.datetime.strptime(s[:10], "%Y-%m-%d")
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            pass
    # dd/mm/aaaa o d/m/aaaa
    if "/" in s:
        partes = s.split("/")
        if len(partes) == 3:
            try:
                return f"{int(partes[0]):02d}/{int(partes[1]):02d}/{partes[2][:4]}"
            except ValueError:
                pass
    return s


def fecha_vacia(v):
    """True si el valor de fecha no representa una fecha real."""
    if v is None:
        return True
    if isinstance(v, float) and pd.isna(v):
        return True
    return str(v).strip() in ("", "nan", "NaT", "None")


def encontrar_xls():
    """Busca el primer .xls en la carpeta actual."""
    archivos = glob.glob("*.xls")
    return archivos[0] if archivos else None


# ── Procesamiento principal ────────────────────────────────────────────────────

def procesar(ruta_entrada):
    # Leer el .xls (formato OLE2/BIFF — requiere xlrd)
    df = pd.read_excel(ruta_entrada, engine="xlrd", header=0)

    # Mapear columnas por nombre normalizado
    col_map = {}
    for col in df.columns:
        cn = norm(col)
        if "fecha" in cn:
            col_map["fecha"] = col
        elif "concepto" in cn or "descripci" in cn:
            col_map["concepto"] = col
        elif "debito" in cn:
            col_map["debito"] = col
        elif "credito" in cn:
            col_map["credito"] = col

    faltantes = [r for r in ("fecha", "concepto", "debito", "credito") if r not in col_map]
    if faltantes:
        raise ValueError(
            f"No se encontraron las columnas requeridas: {faltantes}\n"
            f"Columnas en el archivo: {list(df.columns)}"
        )

    filas = []
    for _, row in df.iterrows():
        # Saltar filas sin fecha (filas vacías o de totales)
        if fecha_vacia(row[col_map["fecha"]]):
            continue

        debe  = to_float(row[col_map["debito"]])
        haber = to_float(row[col_map["credito"]])

        # Saltar filas sin movimiento real (Débito = 0 y Crédito = 0)
        if debe == 0 and haber == 0:
            continue

        fecha    = fmt_fecha(row[col_map["fecha"]])
        concepto = str(row[col_map["concepto"]] or "").strip()

        if haber > 0:
            importe = haber
            tipo    = "Ingreso"
        else:
            importe = debe
            tipo    = "Gasto"

        filas.append((fecha, concepto, importe, tipo))

    if not filas:
        raise ValueError("No se encontraron movimientos válidos en el archivo.")

    # Nombre de salida: mismo nombre + _procesado.xlsx
    base       = os.path.splitext(os.path.abspath(ruta_entrada))[0]
    ruta_salida = base + "_procesado.xlsx"

    # ── Generar XLSX con formato ───────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    # Fila 1: encabezados en negrita
    encabezados = ["Fecha", "Descripción", "Importe", "Tipo"]
    ws.append(encabezados)
    for c in range(1, 5):
        ws.cell(1, c).font = Font(bold=True, name="Calibri", size=11)

    # Datos
    for fecha, desc, importe, tipo in filas:
        ws.append([fecha, desc, importe, tipo])

    # Columna C (Importe): número con 2 decimales, alineado a la derecha
    for fila_n in range(2, len(filas) + 2):
        celda = ws.cell(fila_n, 3)
        celda.number_format = "#,##0.00"
        celda.alignment     = Alignment(horizontal="right")

    # Ancho de columnas
    ws.column_dimensions["A"].width = 13   # Fecha
    ws.column_dimensions["B"].width = 52   # Descripción
    ws.column_dimensions["C"].width = 16   # Importe
    ws.column_dimensions["D"].width = 10   # Tipo

    wb.save(ruta_salida)
    return ruta_salida, len(filas)


# ── Punto de entrada ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) >= 2:
        ruta = sys.argv[1]
    else:
        ruta = encontrar_xls()
        if not ruta:
            print("ERROR: No se encontró ningún archivo .xls en la carpeta actual.")
            print("Uso:   python credicoop_procesador.py [archivo.xls]")
            sys.exit(1)
        print(f"Archivo encontrado: {ruta}")

    if not os.path.isfile(ruta):
        print(f"ERROR: El archivo '{ruta}' no existe.")
        sys.exit(1)

    print(f"Procesando: {ruta} ...")
    try:
        salida, cant = procesar(ruta)
        print(f"Listo → {salida}  ({cant} movimientos)")
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)
